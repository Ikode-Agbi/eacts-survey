from flask import Blueprint, render_template, request, redirect, url_for, flash
from database import db
from data_tables.survey import Survey
from data_tables.question import Question
from data_tables.response import Response
from utils.excel_upload import process_excel_file, check_if_excel_file
from werkzeug.utils import secure_filename
import os

# Create blueprint for admin routes
"""
bluprint groups relted groups together. This one groups all admin pages.
so all routes will have '/admin' in their name
"""
admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

from flask import session

# Add this BEFORE the dashboard route
@admin_bp.route('/login', methods=['GET', 'POST'])
def login():
    """Admin login page."""
    
    if request.method == 'GET':
        return render_template('admin_login.html')
    
    if request.method == 'POST':
        password = request.form.get('password', '')
        
        # TODO: Change this to your actual password
        ADMIN_PASSWORD = 'eacts2026'
        
        if password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            flash('Login successful!', 'success')
            return redirect(url_for('admin.dashboard'))
        else:
            flash('Incorrect password', 'error')
            return redirect(url_for('admin.login'))


@admin_bp.route('/logout')
def logout():
    """Logout admin."""
    session.pop('admin_logged_in', None)
    flash('Logged out successfully', 'success')
    return redirect(url_for('admin.login'))


# Protect all admin routes
@admin_bp.before_request
def check_admin_login():
    """Check if admin is logged in before accessing any admin route."""
    
    # Allow login and logout routes without authentication
    if request.endpoint in ['admin.login', 'admin.logout']:
        return None
    
    # Check if logged in
    if not session.get('admin_logged_in'):
        flash('Please login to access admin panel', 'error')
        return redirect(url_for('admin.login'))


# route 1) admin dashboard 
@admin_bp.route('/')
def dashboard():
    """
   the admin dashboard will show all the surveys

   URL: /admin/
    """
    
    # get all survey and order by created date 
    all_surveys = Survey.query.order_by(Survey.created_at.desc()).all()

    # show the dashboard page with all the surveys
    return render_template('admin_dashboard.html', surveys=all_surveys)


# route 2) upload survey

@admin_bp.route('/upload', methods=['GET', 'POST'])
def upload_survey():
    """Upload Excel file and create survey with sections."""
    
    if request.method == 'GET':
        return render_template('upload_excel.html')
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded', 'error')
            return redirect(request.url)
        
        uploaded_file = request.files['file']
        
        if uploaded_file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)
        
        if not check_if_excel_file(uploaded_file.filename):
            flash('Invalid file type. Please upload Excel (.xlsx or .xls)', 'error')
            return redirect(request.url)
        
        try:
            from flask import current_app
            from data_tables.section import Section
            
            safe_filename = secure_filename(uploaded_file.filename)
            upload_folder = current_app.config['UPLOAD_FOLDER']
            
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder)
            
            temp_file_path = os.path.join(upload_folder, safe_filename)
            uploaded_file.save(temp_file_path)
            
            questions_list = process_excel_file(temp_file_path)
            
            if len(questions_list) == 0:
                flash('No questions found in Excel file', 'error')
                os.remove(temp_file_path)
                return redirect(request.url)
            
            survey_title = request.form.get('title', 'EACTS Consensus Survey')
            survey_description = request.form.get('description', '')
            
            # Create survey
            new_survey = Survey(title=survey_title, description=survey_description)
            db.session.add(new_survey)
            db.session.flush()
            
            # Create one default section
            new_section = Section(
                survey_id=new_survey.id,
                section_number=1,
                title="Questions",
                description=""
            )
            db.session.add(new_section)
            db.session.flush()
            
            # Add all questions to this section
            for index, question_text in enumerate(questions_list):
                new_question = Question(
                    section_id=new_section.id,
                    question_number=index + 1,
                    question_text=question_text
                )
                db.session.add(new_question)
            
            db.session.commit()
            os.remove(temp_file_path)
            
            flash(f'Survey "{survey_title}" created with {len(questions_list)} questions!', 'success')
            return redirect(url_for('admin.dashboard'))
            
        except Exception as error:
            db.session.rollback()
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
            flash(f'Error creating survey: {str(error)}', 'error')
            return redirect(request.url)

@admin_bp.route('/results/<int:survey_id>')
def view_results(survey_id):
    """Show statistics and check which questions meet the 75% threshold."""
    
    survey = Survey.query.get_or_404(survey_id)
    
    # Get all statistics
    all_statistics = survey.get_all_statistics()
    
    # Count total responses
    total_responses = len(survey.responses)
    
    # Count passed/failed questions
    passed_count = sum(1 for stat in all_statistics if stat['meets_threshold'])
    failed_count = sum(1 for stat in all_statistics if not stat['meets_threshold'])
    
    # Get elaborations organized by section and question
    sections_with_elaborations = []
    
    for section in sorted(survey.sections, key=lambda s: s.section_number):
        section_data = {
            'section_title': section.title,
            'questions': []
        }
        
        for question in sorted(section.questions, key=lambda q: q.question_number):
            answers_with_text = []
            
            for answer in question.answers:
                if answer.elaboration and answer.elaboration.strip():
                    answers_with_text.append({
                        'choice': answer.choice,
                        'elaboration': answer.elaboration,
                        'submitted_at': answer.response.submitted_at
                    })
            
            section_data['questions'].append({
                'question_number': question.question_number,
                'question_text': question.question_text,
                'elaborations': answers_with_text
            })
        
        sections_with_elaborations.append(section_data)

    return render_template('view_results.html',
                          survey=survey,
                          stats=all_statistics,
                          total_responses=total_responses,
                          passed_count=passed_count,
                          failed_count=failed_count,
                          sections_with_elaborations=sections_with_elaborations)

@admin_bp.route('/responses/<int:survey_id>')
def view_responses(survey_id):
    """Show all individual responses for a survey."""

    survey = Survey.query.get_or_404(survey_id)

    individual_responses = []
    for resp in sorted(survey.responses, key=lambda r: r.submitted_at, reverse=True):
        display_name = resp.participant_name or 'Anonymous'

        resp_sections = []
        for section in sorted(survey.sections, key=lambda s: s.section_number):
            section_answers = []
            for question in sorted(section.questions, key=lambda q: q.question_number):
                answer = next((a for a in resp.answers if a.question_id == question.id), None)
                section_answers.append({
                    'question_number': question.question_number,
                    'question_text':   question.question_text,
                    'choice':          answer.choice if answer else '—',
                    'elaboration':     answer.elaboration if answer else ''
                })
            resp_sections.append({
                'section_title': section.title,
                'answers': section_answers
            })

        individual_responses.append({
            'id':           resp.id,
            'name':         display_name,
            'submitted_at': resp.submitted_at,
            'is_complete':  resp.is_complete,
            'sections':     resp_sections
        })

    return render_template('individual_responses.html',
                           survey=survey,
                           individual_responses=individual_responses)


@admin_bp.route('/create-manual', methods=['GET', 'POST'])
def create_manual_survey():
    """Create survey manually with sections."""
    
    if request.method == 'GET':
        return render_template('create_manual_survey.html')
    
    if request.method == 'POST':
        try:
            from data_tables.section import Section
            
            survey_title = request.form.get('title')
            survey_description = request.form.get('description', '')
            
            if not survey_title or survey_title.strip() == '':
                flash('Please provide a survey title', 'error')
                return redirect(request.url)
            
            # Create survey
            new_survey = Survey(title=survey_title, description=survey_description)
            db.session.add(new_survey)
            db.session.flush()
            
            # Get sections from form
            section_index = 1
            total_questions = 0
            
            while True:
                section_title_key = f'section_{section_index}_title'
                section_title = request.form.get(section_title_key)
                
                if section_title is None:
                    break
                
                section_title = section_title.strip()
                if section_title == '':
                    section_index += 1
                    continue
                
                section_description = request.form.get(f'section_{section_index}_description', '')
                
                # Create section
                new_section = Section(
                    survey_id=new_survey.id,
                    section_number=section_index,
                    title=section_title,
                    description=section_description
                )
                db.session.add(new_section)
                db.session.flush()
                
                # Get questions for this section
                question_index = 1
                while True:
                    question_key = f'section_{section_index}_question_{question_index}'
                    question_text = request.form.get(question_key)
                    
                    if question_text is None:
                        break
                    
                    question_text = question_text.strip()
                    if question_text != '':
                        new_question = Question(
                            section_id=new_section.id,
                            question_number=question_index,
                            question_text=question_text
                        )
                        db.session.add(new_question)
                        total_questions += 1
                    
                    question_index += 1
                
                section_index += 1
            
            if total_questions == 0:
                flash('Please add at least one question', 'error')
                db.session.rollback()
                return redirect(request.url)
            
            db.session.commit()
            
            flash(f'Survey "{survey_title}" created with {total_questions} questions!', 'success')
            return redirect(url_for('admin.dashboard'))
            
        except Exception as error:
            db.session.rollback()
            flash(f'Error creating survey: {str(error)}', 'error')
            return redirect(request.url)
        

@admin_bp.route('/export-excel/<int:survey_id>')
def export_excel(survey_id):
    """Export survey results to Excel file (questions as rows)."""

    survey = Survey.query.get_or_404(survey_id)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Results'

    # Header row
    headers = ['Q#', 'Section', 'Question Text', 'Total Responses',
               'Yes %', 'No %', 'Abstain', 'Comments']
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Data rows
    for question in survey.get_all_questions():
        stats = question.calculate_statistics()

        # Calculate no percentage (same denominator as yes — excludes abstains)
        total_yes_no = stats['yes_count'] + stats['no_count']
        if total_yes_no > 0:
            no_pct = round((stats['no_count'] / total_yes_no) * 100, 1)
        else:
            no_pct = 0.0

        # Gather comments
        comments = [
            a.elaboration.strip()
            for a in question.answers
            if a.elaboration and a.elaboration.strip()
        ]
        comments_text = ' | '.join(comments) if comments else ''

        ws.append([
            stats['question_number'],
            question.section.title,
            stats['question_text'],
            stats['total_responses'],
            f"{stats['yes_percentage']}%",
            f"{no_pct}%",
            stats['abstain_count'],
            comments_text,
        ])

    # Auto-fit column widths (approximate)
    col_widths = [6, 18, 60, 16, 10, 10, 10, 60]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Wrap text on Question Text and Comments columns
    for row in ws.iter_rows(min_row=2):
        row[2].alignment = Alignment(wrap_text=True, vertical='top')  # Question Text
        row[7].alignment = Alignment(wrap_text=True, vertical='top')  # Comments

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_title = survey.title.replace(' ', '_').replace('/', '_')
    filename = f'{safe_title}_Results.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@admin_bp.route('/export-pdf/<int:survey_id>')
def export_pdf(survey_id):
    """Export survey results to a PDF report."""

    survey = Survey.query.get_or_404(survey_id)

    from flask import send_file
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()

    style_title = ParagraphStyle(
        'SurveyTitle',
        parent=styles['Title'],
        fontSize=20,
        textColor=colors.HexColor('#1B3A5C'),
        spaceAfter=6,
    )
    style_subtitle = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#555555'),
        spaceAfter=16,
    )
    style_section_fail = ParagraphStyle(
        'SectionFail',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#C0392B'),
        spaceBefore=18,
        spaceAfter=6,
    )
    style_section_pass = ParagraphStyle(
        'SectionPass',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#27AE60'),
        spaceBefore=18,
        spaceAfter=6,
    )
    style_question = ParagraphStyle(
        'QuestionText',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#2C3E50'),
        fontName='Helvetica-Bold',
        spaceBefore=10,
        spaceAfter=3,
    )
    style_stats = ParagraphStyle(
        'StatsLine',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#444444'),
        spaceAfter=4,
        leftIndent=12,
    )
    style_comment_label = ParagraphStyle(
        'CommentLabel',
        parent=styles['Normal'],
        fontSize=10,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#555555'),
        spaceBefore=4,
        spaceAfter=2,
        leftIndent=12,
    )
    style_comment = ParagraphStyle(
        'Comment',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#333333'),
        spaceAfter=3,
        leftIndent=24,
        bulletIndent=14,
    )

    # Collect question data
    all_questions = survey.get_all_questions()
    total_responses = len(survey.responses)

    failed_questions = []
    passed_questions = []

    for question in all_questions:
        stats = question.calculate_statistics()
        total_yes_no = stats['yes_count'] + stats['no_count']
        no_pct = round((stats['no_count'] / total_yes_no) * 100, 1) if total_yes_no > 0 else 0.0
        comments = [
            a.elaboration.strip()
            for a in question.answers
            if a.elaboration and a.elaboration.strip()
        ]
        entry = {
            'number': stats['question_number'],
            'text': stats['question_text'],
            'total': stats['total_responses'],
            'yes_pct': stats['yes_percentage'],
            'no_pct': no_pct,
            'abstain': stats['abstain_count'],
            'comments': comments,
        }
        if stats['meets_threshold']:
            passed_questions.append(entry)
        else:
            failed_questions.append(entry)

    passed_count = len(passed_questions)
    failed_count = len(failed_questions)

    story = []

    # Title
    story.append(Paragraph(survey.title, style_title))
    story.append(Paragraph(
        f"Total Responses: {total_responses} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"Passed: {passed_count} &nbsp;&nbsp;|&nbsp;&nbsp; Did Not Pass: {failed_count}",
        style_subtitle
    ))
    story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#DDDDDD'), spaceAfter=10))

    def add_questions(question_list):
        for q in question_list:
            story.append(Paragraph(
                f"Q{q['number']}. {q['text']}",
                style_question
            ))
            story.append(Paragraph(
                f"Respondents: {q['total']} &nbsp;&nbsp;|&nbsp;&nbsp; "
                f"Yes: {q['yes_pct']}% &nbsp;&nbsp;|&nbsp;&nbsp; "
                f"No: {q['no_pct']}% &nbsp;&nbsp;|&nbsp;&nbsp; "
                f"Abstained: {q['abstain']}",
                style_stats
            ))
            if q['comments']:
                story.append(Paragraph('Comments:', style_comment_label))
                for comment in q['comments']:
                    # Escape any HTML special characters in comment text
                    safe_comment = comment.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    story.append(Paragraph(f"\u2022 {safe_comment}", style_comment))
            story.append(Spacer(1, 6))

    # Did Not Pass section
    if failed_questions:
        story.append(Paragraph('DID NOT PASS', style_section_fail))
        story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#E8A0A0'), spaceAfter=6))
        add_questions(failed_questions)

    # Passed section
    if passed_questions:
        story.append(Paragraph('PASSED', style_section_pass))
        story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#A0D8AF'), spaceAfter=6))
        add_questions(passed_questions)

    doc.build(story)
    output.seek(0)

    safe_title = survey.title.replace(' ', '_').replace('/', '_')
    filename = f'{safe_title}_Results.pdf'

    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@admin_bp.route('/toggle/<int:survey_id>', methods=['POST'])
def toggle_survey(survey_id):
    """Toggle a survey between active and inactive."""

    survey = Survey.query.get_or_404(survey_id)

    # Flip the active status
    survey.is_active = not survey.is_active
    db.session.commit()

    status_word = 'activated' if survey.is_active else 'deactivated'
    flash(f'Survey "{survey.title}" has been {status_word}.', 'success')

    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/delete/<int:survey_id>', methods=['POST'])
def delete_survey(survey_id):
    """
    Delete a survey and all its data.
    This removes: survey, questions, responses, and answers.
    """
    
    try:
        # Step 1: Get the survey
        survey = Survey.query.get_or_404(survey_id)
        
        # Step 2: Store the title for the success message
        survey_title = survey.title
        
        # Step 3: Delete it
        # Because of cascade='all, delete-orphan' in our relationships,
        # this automatically deletes all questions, responses, and answers
        db.session.delete(survey)
        db.session.commit()
        
        # Step 4: Show success message
        flash(f'Survey "{survey_title}" deleted successfully', 'success')
        
    except Exception as error:
        # Something went wrong
        db.session.rollback()
        flash(f'Error deleting survey: {str(error)}', 'error')
    
    # Step 5: Go back to admin dashboard
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/delete-response/<int:response_id>', methods=['POST'])
def delete_response(response_id):
    """Delete a single response and its answers (cascade)."""
    try:
        resp = Response.query.get_or_404(response_id)
        survey_id = resp.survey_id
        db.session.delete(resp)
        db.session.commit()
        flash('Response deleted successfully', 'success')
    except Exception as error:
        db.session.rollback()
        flash(f'Error deleting response: {str(error)}', 'error')
    return redirect(url_for('admin.individual_responses', survey_id=survey_id))


@admin_bp.route('/edit/<int:survey_id>')
def edit_survey(survey_id):
    """Edit survey - organize questions into sections."""
    
    survey = Survey.query.get_or_404(survey_id)
    
    # Get sections in order
    sections = sorted(survey.sections, key=lambda s: s.section_number)
    
    return render_template('edit_survey.html', survey=survey, sections=sections)


@admin_bp.route('/edit/<int:survey_id>/update', methods=['POST'])
def update_survey(survey_id):
    """Process survey edits."""
    
    from data_tables.section import Section
    
    survey = Survey.query.get_or_404(survey_id)
    
    try:
        # Update survey title and description
        survey.title = request.form.get('title', survey.title)
        survey.description = request.form.get('description', '')
        
        # Delete all existing sections and questions
        # We'll recreate them from the form
        for section in survey.sections:
            db.session.delete(section)
        
        db.session.flush()
        
        # Rebuild sections from form
        section_index = 1
        
        while True:
            section_title_key = f'section_{section_index}_title'
            section_title = request.form.get(section_title_key)
            
            if section_title is None:
                break
            
            section_description = request.form.get(f'section_{section_index}_description', '')
            
            # Create section
            new_section = Section(
                survey_id=survey.id,
                section_number=section_index,
                title=section_title,
                description=section_description
            )
            db.session.add(new_section)
            db.session.flush()
            
            # Add questions for this section
            question_index = 1
            while True:
                question_key = f'section_{section_index}_question_{question_index}'
                question_text = request.form.get(question_key)
                
                if question_text is None:
                    break
                
                question_text = question_text.strip()
                if question_text:
                    new_question = Question(
                        section_id=new_section.id,
                        question_number=question_index,
                        question_text=question_text
                    )
                    db.session.add(new_question)
                
                question_index += 1
            
            section_index += 1
        
        db.session.commit()
        flash('Survey updated successfully!', 'success')
        return redirect(url_for('admin.dashboard'))
        
    except Exception as error:
        db.session.rollback()
        flash(f'Error updating survey: {str(error)}', 'error')
        return redirect(url_for('admin.edit_survey', survey_id=survey_id))